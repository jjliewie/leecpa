import os
from openpyxl import load_workbook
from copy import copy

humber = 'C:/Users/Julie/Desktop/hv/humber.xlsx'
valleywood = 'C:/Users/Julie/Desktop/hv/valleywood.xlsx'

hbook = load_workbook(filename=humber)
vbook = load_workbook(filename=valleywood)

def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)
    for key, _ in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width)
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

# from oscar on stackoverflow
# https://stackoverflow.com/questions/42344041/how-to-copy-worksheet-from-one-workbook-to-another-one-using-openpyxl

months = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]

def priority(wb):
    month = 12
    day = 31
    for i in range(len(months)):
        if months[i] in wb.lower():
            month = i+1
            break
    for i in range(len(wb)):
        if wb[i].isdigit():
            if i < len(wb) - 1 and wb[i+1].isdigit():
                day = int(wb[i] + wb[i+1])
                break
            day = int(wb[i])
            break
    return month + (day / 100)

def name(wb):
    start = 0
    for i in range(len(wb) - 3):
        k = wb[i] + wb[i+1] + wb[i+2]
        if k.lower() in months:
            start = i
            break
    end = len(wb)
    for i in range(start, len(wb)):
        if wb[i] == '.':
            end = i
            break
    return wb[start:end]

def ordering(lst):
    pdict = {}
    for i in lst:
        pdict[priority(i)] = i
    return list(dict(sorted(pdict.items())).values())

def makesheets(book, name):
    hmain = book.worksheets[0]
    vmain = book.worksheets[1]
    hsheet = hbook.create_sheet(name)
    vsheet = vbook.create_sheet(name)
    copy_sheet(hmain, hsheet)
    copy_sheet(vmain, vsheet)
    hbook.save(humber)
    vbook.save(valleywood)

path = 'C:/Users/Julie/Desktop/hv/all'
spreadsheets = ordering(os.listdir(path))

for wb in spreadsheets:
    specific = path + '/' + wb
    book = load_workbook(filename=specific)
    makesheets(book, name(wb))

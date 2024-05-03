import os
import openpyxl as xl
from openpyxl import load_workbook

humber = 'C:/Users/Julie/Desktop/hv/humber.xlsx'
valleywood = 'C:/Users/Julie/Desktop/hv/valleywood.xlsx'

hbook = load_workbook(filename=humber)
vbook = load_workbook(filename=valleywood)

months = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]

def priority(wb):
    month = 12
    day = 31
    for i in range(len(months)):
        if months[i] in wb:
            month = i+1
            break
    for i in range(len(wb)):
        if wb[i].isdigit():
            if i < len(wb) - 1 and wb[i+1].isdigit():
                day = int(wb[i] + wb[i+1])
                break
            day = int(wb[i])
            break
    return month + (day / 100)

def name(wb):
    start = 0
    for i in range(len(wb) - 3):
        k = wb[i] + wb[i+1] + wb[i+2]
        if k in months:
            start = i
            break
    return wb[i:]

def ordering(lst):
    pdict = {}
    for i in lst:
        pdict[priority(i)] = i
    return list(dict(sorted(pdict.items())).values())

def copysheet(book, name):
    hmain = book.worksheets[0]
    vmain = book.worksheets[1]
    hsheet = hbook.create_sheet(name)
    vsheet = vbook.create_sheet(name)

    for row in hmain:
        for cell in row:
            hsheet[cell.coordinate].value = cell.value

    for row in vmain:
        for cell in row:
            vsheet[cell.coordinate].value = cell.value

    hbook.save(humber)
    vbook.save(valleywood)

path = 'C:/Users/Julie/Desktop/hv/all'
spreadsheets = ordering(os.listdir(path))
# print(spreadsheets)

for wb in spreadsheets:
    specific = path + '/' + wb
    book = load_workbook(filename=specific)
    copysheet(book, name(wb))
